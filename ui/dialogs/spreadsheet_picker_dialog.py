"""
Spreadsheet Picker Dialog

Pick mode:  merged/reordered card list showing count of identical rows.
Edit mode:  full QTableWidget with column × delete, column/row drag-to-reorder,
            auto-grow empty row, and save back to the original file.
"""

import csv
import os

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QScrollArea, QWidget,
    QPushButton, QLabel, QSizePolicy, QFrame, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QAbstractItemDelegate,
    QInputDialog, QMessageBox, QStackedWidget, QComboBox, QStyledItemDelegate,
    QMenu,
)
from PySide6.QtCore import Qt, Signal, QRect, QTimer
from PySide6.QtGui import QFont, QColor


# ---------------------------------------------------------------------------
# File I/O
# ---------------------------------------------------------------------------

def read_spreadsheet(path: str) -> list[list[str]]:
    """Return all rows as lists of strings. Row 0 is the header."""
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [row for row in csv.reader(f)]

    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl not installed — run: pip install openpyxl")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        return [["" if v is None else str(v) for v in row]
                for row in ws.iter_rows(values_only=True)]

    if ext == ".ods":
        try:
            from odf.opendocument import load as ods_load
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
        except ImportError:
            raise ValueError("odfpy not installed — run: pip install odfpy")
        doc = ods_load(path)
        rows = []
        for sheet in doc.spreadsheet.getElementsByType(Table):
            for tr in sheet.getElementsByType(TableRow):
                cells = tr.getElementsByType(TableCell)
                row = []
                for cell in cells:
                    repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                    ps = cell.getElementsByType(P)
                    text = "".join(str(p) for p in ps) if ps else ""
                    row.extend([text] * repeat)
                rows.append(row)
            break
        return rows

    if ext == ".xls":
        try:
            import xlrd
        except ImportError:
            raise ValueError("xlrd not installed — run: pip install xlrd")
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        return [[str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                for r in range(ws.nrows)]

    raise ValueError(f"Unsupported format: {ext}")


def write_spreadsheet(path: str, headers: list[str], rows: list[list[str]]):
    """Write headers + rows to a file, overwriting it."""
    all_rows = [headers] + rows
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(all_rows)
        return

    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in all_rows:
            ws.append(row)
        wb.save(path)
        return

    if ext == ".ods":
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        doc = OpenDocumentSpreadsheet()
        table = Table(name="Sheet1")
        for row in all_rows:
            tr = TableRow()
            for val in row:
                tc = TableCell()
                tc.addElement(P(text=str(val)))
                tr.addElement(tc)
            table.addElement(tr)
        doc.spreadsheet.addElement(table)
        doc.save(path)
        return

    raise ValueError(f"Unsupported format for writing: {ext}")


# ---------------------------------------------------------------------------
# Delegate for "side" column — G / D combo box
# ---------------------------------------------------------------------------

class SideDelegate(QStyledItemDelegate):
    """Shows a G / D combo box when editing any cell in the 'side' column.
    The dropdown opens automatically on the second click (no 3rd click needed)."""

    _OPTIONS = ["", "G", "D"]
    _STYLE = """
        QComboBox {
            background-color: #1d1f28;
            color: #ffffff;
            border: 1px solid #BB86FC;
            padding: 2px 6px;
        }
        QComboBox::drop-down { border: none; width: 18px; }
        QComboBox QAbstractItemView {
            background-color: #282a36;
            color: #ffffff;
            selection-background-color: #44475a;
            border: 1px solid #44475a;
        }
    """

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(self._OPTIONS)
        combo.setStyleSheet(self._STYLE)
        # activated fires only on user selection (not programmatic), avoiding the
        # "commitData called with editor that does not belong" warning
        combo.activated.connect(lambda _: self._commit_close(combo))
        # Auto-open the dropdown as soon as the editor is shown
        QTimer.singleShot(0, combo.showPopup)
        return combo

    def _commit_close(self, combo):
        self.commitData.emit(combo)
        self.closeEditor.emit(combo, QAbstractItemDelegate.NoHint)

    def setEditorData(self, editor, index):
        value = index.data(Qt.EditRole) or ""
        editor.blockSignals(True)
        idx = editor.findText(value, Qt.MatchFixedString)
        editor.setCurrentIndex(idx if idx >= 0 else 0)
        editor.blockSignals(False)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


# ---------------------------------------------------------------------------
# Custom column header — × delete button + drag-to-reorder
# ---------------------------------------------------------------------------

class ColumnHeader(QHeaderView):
    """Horizontal header that paints a × button on each section."""

    column_delete_requested = Signal(int)   # emits logical index

    _BTN_W = 16
    _BTN_H = 16
    _BTN_PAD = 4

    def __init__(self, parent=None):
        super().__init__(Qt.Horizontal, parent)
        self.setSectionsMovable(True)
        self.setHighlightSections(False)
        self.setSectionResizeMode(QHeaderView.Interactive)
        self.setMinimumSectionSize(80)
        self.setDefaultSectionSize(130)

    def paintSection(self, painter, rect, logical_index):
        super().paintSection(painter, rect, logical_index)
        btn = self._btn_rect(rect)
        painter.save()
        painter.setPen(QColor("#ff5555"))
        painter.setFont(QFont("Arial", 11, QFont.Bold))
        painter.drawText(btn, Qt.AlignCenter, "✕")
        painter.restore()

    def _btn_rect(self, section_rect: QRect) -> QRect:
        x = section_rect.right() - self._BTN_W - self._BTN_PAD
        y = section_rect.top() + (section_rect.height() - self._BTN_H) // 2
        return QRect(x, y, self._BTN_W, self._BTN_H)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            logical = self.logicalIndexAt(event.pos())
            if logical >= 0:
                vpos = self.sectionViewportPosition(logical)
                w = self.sectionSize(logical)
                sec_rect = QRect(vpos, 0, w, self.height())
                if self._btn_rect(sec_rect).contains(event.pos()):
                    self.column_delete_requested.emit(logical)
                    return
        super().mousePressEvent(event)


# ---------------------------------------------------------------------------
# Row card (pick mode)
# ---------------------------------------------------------------------------

# Columns shown in pick view, in this order (others appended after)
PICK_DISPLAY_COLS = ["frame_height", "side", "frame_width", "frame total width"]


class RowCard(QFrame):
    STATE_DEFAULT  = "default"
    STATE_LAST     = "last"
    STATE_SELECTED = "selected"

    def __init__(self, display_number: int, original_index: int,
                 count: int, values: list[str], parent=None):
        super().__init__(parent)
        self.original_index = original_index
        self._state    = self.STATE_DEFAULT
        self._on_click = None

        self.setFrameShape(QFrame.StyledPanel)
        self.setCursor(Qt.PointingHandCursor)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(8, 6, 8, 6)
        layout.setSpacing(12)

        # Row number badge
        num = QLabel(f"#{display_number}")
        num.setFixedWidth(36)
        num.setAlignment(Qt.AlignCenter)
        num.setFont(QFont("Consolas", 12, QFont.Bold))
        num.setStyleSheet("color: #6f779a;")
        layout.addWidget(num)

        # Count badge
        count_lbl = QLabel(f"×{count}")
        count_lbl.setFixedWidth(32)
        count_lbl.setAlignment(Qt.AlignCenter)
        count_lbl.setFont(QFont("Consolas", 10, QFont.Bold))
        count_lbl.setStyleSheet("color: #23c87b;" if count > 1 else "color: #44475a;")
        layout.addWidget(count_lbl)

        # Value columns
        for val in values:
            lbl = QLabel(val if val else "—")
            lbl.setFont(QFont("Consolas", 11))
            lbl.setStyleSheet("color: #ffffff;")
            layout.addWidget(lbl, 1)

        self._apply_style()

    def set_state(self, state: str):
        self._state = state
        self._apply_style()

    def _apply_style(self):
        colors = {
            self.STATE_DEFAULT:  "#44475a",
            self.STATE_LAST:     "#4a9eff",
            self.STATE_SELECTED: "#23c87b",
        }
        border = colors.get(self._state, "#44475a")
        self.setStyleSheet(f"""
            RowCard {{
                background-color: transparent;
                border: 2px solid {border};
                border-radius: 5px;
                margin: 2px 0;
            }}
            RowCard:hover {{ background-color: #3a3c4e; }}
        """)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and self._on_click:
            self._on_click(self.original_index)
        super().mousePressEvent(event)

    def set_click_handler(self, fn):
        self._on_click = fn


# ---------------------------------------------------------------------------
# Main dialog
# ---------------------------------------------------------------------------

class SpreadsheetPickerDialog(QDialog):

    def __init__(self, file_path: str, known_variables: dict,
                 last_row_index: int | None, parent=None,
                 start_in_edit_mode: bool = False):
        super().__init__(parent)
        self.setWindowTitle("Spreadsheet")
        self.setModal(True)
        self.resize(1050, 620)

        self._file_path       = file_path
        self._known_variables = known_variables
        self._last_row_index  = last_row_index
        self._cards: list[RowCard] = []
        self._all_rows: list[list[str]] = []
        self._col_map: dict[int, str] = {}
        self._merged_groups: list = []        # [count, original_index, row_tuple]
        self._selected_merged_idx: int | None = None
        self._in_edit_mode = False
        self._dirty        = False

        # Dialog outputs
        self.picked_row_index: int | None  = None
        self.picked_count: int             = 1
        self.picked_display_data: dict     = {}
        self.values_to_apply: dict         = {}

        self._apply_style()
        self._build_ui()

        try:
            self._all_rows = read_spreadsheet(file_path)
        except Exception as exc:
            self._show_error(str(exc))

        self._populate_pick_view()

        if start_in_edit_mode:
            self._enter_edit_mode()

    # ------------------------------------------------------------------ style

    def _apply_style(self):
        self.setStyleSheet("""
            SpreadsheetPickerDialog { background-color: #282a36; color: #ffffff; }
            QLabel { color: #ffffff; background-color: transparent; }
            QScrollArea {
                background-color: #1d1f28;
                border: 1px solid #44475a;
                border-radius: 4px;
            }
            QTableWidget {
                background-color: #1d1f28;
                color: #ffffff;
                border: 1px solid #44475a;
                gridline-color: #3a3c4e;
                selection-background-color: #44475a;
            }
            QTableWidget::item { padding: 4px 6px; }
            QTableWidget::item:selected { background-color: #44475a; }
            QHeaderView::section {
                background-color: #282a36;
                color: #BB86FC;
                border: 1px solid #44475a;
                padding: 5px 28px 5px 8px;
                font-family: Consolas; font-size: 11px; font-weight: bold;
            }
            QTableCornerButton::section { background-color: #282a36; }
            QScrollBar:vertical { background: #1d1f28; width: 10px; }
            QScrollBar::handle:vertical { background: #44475a; border-radius: 5px; min-height: 20px; }
            QScrollBar:horizontal { background: #1d1f28; height: 10px; }
            QScrollBar::handle:horizontal { background: #44475a; border-radius: 5px; min-width: 20px; }
            QInputDialog { background-color: #282a36; color: #ffffff; }
        """)

    # ------------------------------------------------------------------ build UI

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setSpacing(8)
        root.setContentsMargins(12, 12, 12, 12)

        # ── Title row ──────────────────────────────────────────────────
        title_row = QHBoxLayout()
        title_lbl = QLabel("Spreadsheet")
        title_lbl.setFont(QFont("Arial", 13, QFont.Bold))
        title_row.addWidget(title_lbl)

        hint = QLabel(os.path.basename(self._file_path))
        hint.setFont(QFont("Consolas", 9))
        hint.setStyleSheet("color: #6f779a;")
        hint.setToolTip(self._file_path)
        title_row.addWidget(hint)
        title_row.addStretch()

        self._edit_btn = self._mk_btn("Edit Table", "#BB86FC", checkable=True)
        self._edit_btn.toggled.connect(self._on_edit_toggled)
        title_row.addWidget(self._edit_btn)
        root.addLayout(title_row)

        # ── Error label ────────────────────────────────────────────────
        self._error_lbl = QLabel()
        self._error_lbl.setStyleSheet("color: #ff5555;")
        self._error_lbl.setVisible(False)
        root.addWidget(self._error_lbl)

        # ── Edit toolbar (hidden in pick mode) ─────────────────────────
        self._edit_toolbar = QWidget()
        tb = QHBoxLayout(self._edit_toolbar)
        tb.setContentsMargins(0, 0, 0, 0)
        tb.setSpacing(6)
        add_col_btn = self._mk_btn("+ Add Column", "#23c87b")
        add_col_btn.clicked.connect(self._add_column)
        tb.addWidget(add_col_btn)
        tb.addStretch()
        self._edit_toolbar.setVisible(False)
        root.addWidget(self._edit_toolbar)

        # ── Stacked widget ─────────────────────────────────────────────
        self._stack = QStackedWidget()
        root.addWidget(self._stack, 1)

        # Page 0 — pick view
        pick_page = QWidget()
        pp_layout = QVBoxLayout(pick_page)
        pp_layout.setContentsMargins(0, 0, 0, 0)
        pick_scroll = QScrollArea()
        pick_scroll.setWidgetResizable(True)
        self._pick_container = QWidget()
        self._pick_container.setStyleSheet("background-color: #1d1f28;")
        self._rows_layout = QVBoxLayout(self._pick_container)
        self._rows_layout.setContentsMargins(6, 6, 6, 6)
        self._rows_layout.setSpacing(4)
        pick_scroll.setWidget(self._pick_container)
        pp_layout.addWidget(pick_scroll)
        self._stack.addWidget(pick_page)

        # Page 1 — edit view
        edit_page = QWidget()
        ep_layout = QVBoxLayout(edit_page)
        ep_layout.setContentsMargins(0, 0, 0, 0)
        self._table = QTableWidget()
        self._col_header = ColumnHeader(self._table)
        self._table.setHorizontalHeader(self._col_header)
        self._col_header.column_delete_requested.connect(self._delete_column)
        self._table.verticalHeader().setSectionsMovable(True)
        self._table.setEditTriggers(
            QAbstractItemView.AnyKeyPressed |
            QAbstractItemView.DoubleClicked |
            QAbstractItemView.SelectedClicked
        )
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.itemChanged.connect(self._on_item_changed)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_table_context_menu)
        self._table.verticalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.verticalHeader().customContextMenuRequested.connect(
            lambda pos: self._on_table_context_menu(
                self._table.verticalHeader().mapTo(self._table.viewport(), pos)
            )
        )
        ep_layout.addWidget(self._table)
        self._stack.addWidget(edit_page)

        # ── Bottom buttons ─────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.addStretch()

        cancel_btn = self._mk_btn("Cancel", "#BB86FC")
        cancel_btn.clicked.connect(self.reject)

        self._confirm_btn = self._mk_btn("Confirm", "#23c87b")
        self._confirm_btn.setEnabled(False)
        self._confirm_btn.clicked.connect(self._on_confirm)

        self._save_btn = self._mk_btn("Save Changes", "#23c87b")
        self._save_btn.setVisible(False)
        self._save_btn.clicked.connect(self._on_save)

        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(self._confirm_btn)
        btn_row.addWidget(self._save_btn)
        root.addLayout(btn_row)

    @staticmethod
    def _mk_btn(text: str, color: str, checkable: bool = False) -> QPushButton:
        btn = QPushButton(text)
        btn.setCheckable(checkable)
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #1d1f28; color: {color};
                border: 2px solid {color}; border-radius: 4px;
                padding: 6px 14px; min-width: 80px;
            }}
            QPushButton:hover {{ background-color: #0d0f18; }}
            QPushButton:pressed, QPushButton:checked {{
                background-color: {color}; color: #1d1f28;
            }}
            QPushButton:disabled {{ color: #6f779a; border-color: #6f779a; }}
        """)
        return btn

    # ------------------------------------------------------------------ pick view

    def _populate_pick_view(self):
        while self._rows_layout.count():
            item = self._rows_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._cards.clear()
        self._merged_groups.clear()
        self._selected_merged_idx = None
        self._col_map.clear()
        if hasattr(self, '_confirm_btn'):
            self._confirm_btn.setEnabled(False)

        if not self._all_rows:
            return

        width = max(len(r) for r in self._all_rows)
        rows  = [r + [""] * (width - len(r)) for r in self._all_rows]
        header    = rows[0]
        data_rows = rows[1:]

        # Build col_map: column index → variable name
        for i, h in enumerate(header):
            clean = h.strip().lstrip("$")
            if clean and clean in self._known_variables:
                self._col_map[i] = clean

        if not data_rows:
            self._show_error("No data rows found (header only). Use Edit Table to add rows.")
            return

        # --- Build display column order: PICK_DISPLAY_COLS first, then remaining ---
        pick_col_indices: list[int] = []
        pick_col_labels: list[str]  = []
        for name in PICK_DISPLAY_COLS:
            for i, h in enumerate(header):
                if h.strip().lower() == name.lower():
                    pick_col_indices.append(i)
                    pick_col_labels.append(h.strip())
                    break

        remaining_indices = [i for i in range(len(header)) if i not in pick_col_indices]
        remaining_labels  = [header[i].strip() for i in remaining_indices]

        all_display_indices = pick_col_indices + remaining_indices
        all_display_labels  = pick_col_labels  + remaining_labels

        # --- Merge identical rows (same full content) ---
        seen: dict[tuple, int] = {}   # row_tuple → index in _merged_groups
        for orig_idx, row in enumerate(data_rows):
            key = tuple(row)
            if key in seen:
                self._merged_groups[seen[key]][0] += 1
            else:
                seen[key] = len(self._merged_groups)
                self._merged_groups.append([1, orig_idx, key])

        # --- Header strip ---
        hdr_frame = QFrame()
        hdr_frame.setStyleSheet("background-color: #282a36; border: none;")
        h_layout = QHBoxLayout(hdr_frame)
        h_layout.setContentsMargins(8, 4, 8, 4)
        h_layout.setSpacing(12)

        # Row # spacer
        spacer = QLabel("")
        spacer.setFixedWidth(36)
        h_layout.addWidget(spacer)

        # Count header
        cnt_hdr = QLabel("×")
        cnt_hdr.setFixedWidth(32)
        cnt_hdr.setAlignment(Qt.AlignCenter)
        cnt_hdr.setFont(QFont("Consolas", 10, QFont.Bold))
        cnt_hdr.setStyleSheet("color: #BB86FC;")
        h_layout.addWidget(cnt_hdr)

        for label in all_display_labels:
            lbl = QLabel(label)
            lbl.setFont(QFont("Consolas", 10, QFont.Bold))
            lbl.setStyleSheet("color: #BB86FC;")
            h_layout.addWidget(lbl, 1)

        self._rows_layout.addWidget(hdr_frame)

        # --- Cards ---
        for disp_num, (count, orig_idx, row_tuple) in enumerate(self._merged_groups, 1):
            row = list(row_tuple)
            display_values = [row[i] if i < len(row) else "" for i in all_display_indices]
            card = RowCard(disp_num, orig_idx, count, display_values)
            card.set_click_handler(self._on_card_clicked)
            self._cards.append(card)
            self._rows_layout.addWidget(card)

        self._rows_layout.addStretch()

        # Mark last-confirmed card as blue
        if self._last_row_index is not None:
            for card in self._cards:
                if card.original_index == self._last_row_index:
                    card.set_state(RowCard.STATE_LAST)
                    break

    def _on_card_clicked(self, original_index: int):
        for i, card in enumerate(self._cards):
            if card.original_index == original_index:
                card.set_state(RowCard.STATE_SELECTED)
                self._selected_merged_idx = i
            elif self._last_row_index is not None and card.original_index == self._last_row_index:
                card.set_state(RowCard.STATE_LAST)
            else:
                card.set_state(RowCard.STATE_DEFAULT)
        self._confirm_btn.setEnabled(True)

    def _on_confirm(self):
        if self._selected_merged_idx is None:
            return
        count, orig_idx, row_tuple = self._merged_groups[self._selected_merged_idx]
        row = list(row_tuple)

        # Build values_to_apply from col_map
        result = {}
        for col_i, var_name in self._col_map.items():
            if col_i < len(row):
                val = row[col_i].strip()
                if val:
                    result[var_name] = val

        self.values_to_apply  = result
        self.picked_row_index = orig_idx
        self.picked_count     = count

        # Build display data for the tab info panel
        header = self._all_rows[0] if self._all_rows else []
        self.picked_display_data = {"count": count}
        for name in PICK_DISPLAY_COLS:
            for i, h in enumerate(header):
                if h.strip().lower() == name.lower() and i < len(row):
                    self.picked_display_data[name] = row[i]
                    break

        self.accept()

    # ------------------------------------------------------------------ edit mode

    def _on_edit_toggled(self, checked: bool):
        if checked:
            self._enter_edit_mode()
        else:
            self._leave_edit_mode()

    def _enter_edit_mode(self):
        self._in_edit_mode = True
        self._dirty = False
        self._error_lbl.setVisible(False)
        self._populate_table()
        self._stack.setCurrentIndex(1)
        self._edit_toolbar.setVisible(True)
        self._confirm_btn.setVisible(False)
        self._save_btn.setVisible(True)
        self._edit_btn.setText("← Pick Mode")
        self._edit_btn.setChecked(True)

    def _leave_edit_mode(self):
        if self._dirty:
            ans = QMessageBox.question(
                self, "Unsaved Changes",
                "You have unsaved changes. Save before leaving edit mode?",
                QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
            )
            if ans == QMessageBox.Cancel:
                self._edit_btn.setChecked(True)
                return
            if ans == QMessageBox.Save and not self._do_save():
                self._edit_btn.setChecked(True)
                return

        self._in_edit_mode = False
        self._dirty = False
        self._stack.setCurrentIndex(0)
        self._edit_toolbar.setVisible(False)
        self._confirm_btn.setVisible(True)
        self._save_btn.setVisible(False)
        self._edit_btn.setText("Edit Table")
        self._edit_btn.setChecked(False)

    # ------------------------------------------------------------------ table population

    def _populate_table(self):
        self._table.blockSignals(True)

        if not self._all_rows:
            self._table.setRowCount(1)
            self._table.setColumnCount(0)
            self._table.blockSignals(False)
            return

        header    = self._all_rows[0]
        data_rows = self._all_rows[1:]

        self._table.setColumnCount(len(header))
        self._table.setRowCount(len(data_rows) + 1)   # +1 for trailing empty row
        self._table.setHorizontalHeaderLabels(header)
        self._table.horizontalHeader().setStretchLastSection(True)

        for r, row in enumerate(data_rows):
            for c in range(len(header)):
                val = row[c] if c < len(row) else ""
                self._table.setItem(r, c, QTableWidgetItem(val))

        # Trailing empty row
        empty_r = len(data_rows)
        for c in range(len(header)):
            self._table.setItem(empty_r, c, QTableWidgetItem(""))

        # Apply G/D delegate to the "side" column (if present)
        for c, col_name in enumerate(header):
            if col_name.strip().lower() == "side":
                self._table.setItemDelegateForColumn(c, SideDelegate(self._table))
                break

        self._table.blockSignals(False)

    def _get_table_data(self) -> tuple[list[str], list[list[str]]]:
        """Read table data in the current visual order (rows and columns may be reordered)."""
        h_hdr = self._table.horizontalHeader()
        v_hdr = self._table.verticalHeader()
        nc    = self._table.columnCount()
        nr    = self._table.rowCount()

        headers = []
        for vc in range(nc):
            lc   = h_hdr.logicalIndex(vc)
            item = self._table.horizontalHeaderItem(lc)
            headers.append(item.text() if item else "")

        rows = []
        for vr in range(nr):
            lr  = v_hdr.logicalIndex(vr)
            row = []
            for vc in range(nc):
                lc   = h_hdr.logicalIndex(vc)
                item = self._table.item(lr, lc)
                row.append(item.text() if item else "")
            rows.append(row)

        return headers, rows

    def _on_item_changed(self, item):
        self._dirty = True
        # Auto-add a trailing empty row when the user fills the last row
        nr = self._table.rowCount()
        if nr > 0 and item.row() == nr - 1:
            nc = self._table.columnCount()
            has_content = any(
                self._table.item(nr - 1, c) and bool(self._table.item(nr - 1, c).text().strip())
                for c in range(nc)
            )
            if has_content:
                self._table.blockSignals(True)
                self._table.insertRow(nr)
                for c in range(nc):
                    self._table.setItem(nr, c, QTableWidgetItem(""))
                self._table.blockSignals(False)

    # ------------------------------------------------------------------ save

    def _do_save(self) -> bool:
        headers, rows = self._get_table_data()
        # Strip trailing empty rows before writing
        while rows and not any(v.strip() for v in rows[-1]):
            rows.pop()
        try:
            write_spreadsheet(self._file_path, headers, rows)
            self._all_rows = [headers] + rows
            self._populate_pick_view()
            self._dirty = False
            return True
        except Exception as exc:
            QMessageBox.critical(self, "Save Error", str(exc))
            return False

    def _on_save(self):
        if self._do_save():
            self._leave_edit_mode()

    # ------------------------------------------------------------------ edit actions

    def _add_column(self):
        existing = set()
        for c in range(self._table.columnCount()):
            item = self._table.horizontalHeaderItem(c)
            if item:
                existing.add(item.text().strip().lstrip("$"))

        suggestions = [k for k in self._known_variables if k not in existing]

        name, ok = QInputDialog.getItem(
            self, "Add Column", "Variable name:",
            suggestions, 0, True,
        )
        if not ok or not name.strip():
            return

        name = name.strip().lstrip("$")
        nc   = self._table.columnCount()
        self._table.blockSignals(True)
        self._table.insertColumn(nc)
        self._table.setHorizontalHeaderItem(nc, QTableWidgetItem(name))
        for r in range(self._table.rowCount()):
            self._table.setItem(r, nc, QTableWidgetItem(""))
        self._table.blockSignals(False)
        self._dirty = True

    def _delete_column(self, logical_index: int):
        hdr  = self._table.horizontalHeaderItem(logical_index)
        name = hdr.text() if hdr else f"Column {logical_index + 1}"
        ans  = QMessageBox.question(
            self, "Delete Column",
            f'Delete column "{name}"?',
            QMessageBox.Yes | QMessageBox.No,
        )
        if ans == QMessageBox.Yes:
            self._table.removeColumn(logical_index)

    # ------------------------------------------------------------------ row context menu

    _MENU_STYLE = """
        QMenu {
            background-color: #1d1f28;
            color: #ffffff;
            border: 1px solid #44475a;
            border-radius: 4px;
        }
        QMenu::item { padding: 6px 18px; }
        QMenu::item:selected { background-color: #44475a; }
        QMenu::separator { height: 1px; background: #44475a; margin: 3px 0; }
    """

    def _on_table_context_menu(self, pos):
        row = self._table.rowAt(pos.y())
        if row < 0:
            return
        menu = QMenu(self)
        menu.setStyleSheet(self._MENU_STYLE)
        menu.addAction("Duplicate Row",     lambda: self._duplicate_row(row))
        menu.addSeparator()
        menu.addAction("Insert Row Above",  lambda: self._insert_row_at(row))
        menu.addAction("Insert Row Below",  lambda: self._insert_row_at(row + 1))
        menu.addSeparator()
        menu.addAction("Delete Row",        lambda: self._delete_row(row))
        menu.exec(self._table.viewport().mapToGlobal(pos))

    def _delete_row(self, row: int):
        self._table.removeRow(row)
        self._dirty = True
        # Always keep a trailing empty row
        nr = self._table.rowCount()
        nc = self._table.columnCount()
        needs_empty = nr == 0 or any(
            self._table.item(nr - 1, c) and self._table.item(nr - 1, c).text().strip()
            for c in range(nc)
        )
        if needs_empty:
            self._table.blockSignals(True)
            self._table.insertRow(nr)
            for c in range(nc):
                self._table.setItem(nr, c, QTableWidgetItem(""))
            self._table.blockSignals(False)

    def _duplicate_row(self, row: int):
        nc = self._table.columnCount()
        self._table.blockSignals(True)
        self._table.insertRow(row + 1)
        for c in range(nc):
            src = self._table.item(row, c)
            self._table.setItem(row + 1, c, QTableWidgetItem(src.text() if src else ""))
        self._table.blockSignals(False)
        self._dirty = True

    def _insert_row_at(self, row: int):
        self._table.blockSignals(True)
        self._table.insertRow(row)
        for c in range(self._table.columnCount()):
            self._table.setItem(row, c, QTableWidgetItem(""))
        self._table.blockSignals(False)
        self._dirty = True

    # ------------------------------------------------------------------ helpers

    def _show_error(self, msg: str):
        self._error_lbl.setText(f"⚠  {msg}")
        self._error_lbl.setVisible(True)
